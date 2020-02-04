from datetime import datetime
from msilib.schema import Font

from openpyxl import Workbook
from openpyxl.drawing import colors
from openpyxl.drawing.image import Image


class ExcelUtils(object):
    """
    pip install openpyxl
    pip install pillow
    """

    def __init__(self):
        self.wb = Workbook()  # 增加一个工作表的对象
        self.ws = self.wb.active  # 默认激活哟个表单
        self.ws_two = self.wb.create_sheet('我的表单')  # 指定标题
        self.ws.title = '你的表单'
        self.ws.sheet_properties.tabColor = 'ff0000'
        self.ws_three = self.wb.create_sheet()

    def do_sth(self):
        # 插入数据
        self.ws['A1'] = 66
        self.ws['A2'] = '你好'
        self.ws['A3'] = datetime.now()
        self.wb.save('./static/text.xlsx')

        for row in self.ws_two['A1:E5']:
            for cell in row:
                cell.value = '你好'
        # 插入图片
        img = Image('./static/lay.jpg')
        self.ws.add_image(img, 'B1')

        # 设置文字
        # font = Font(sz=18, color)
        # self.ws['A2'].font=font
        # 合并单元格
        self.ws.merge_cells('A4:E5')

        self.wb.save('./static/test.xlsx')


if __name__ == '__main__':
    client = ExcelUtils()
    client.do_sth()
